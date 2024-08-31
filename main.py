import os
from PIL import Image, ImageGrab
import io
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog

def resize_image(image, width=1000):
    aspect_ratio = image.height / image.width
    new_height = int(width * aspect_ratio)
    return image.resize((width, new_height))

def create_excel_report(test_name, test_number, steps, images, result, defect=None, export_path=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Report"

    for i, (step, image) in enumerate(zip(steps, images), start=1):
        ws.cell(row=i*3-2, column=1, value=f"{step}")
        ws.cell(row=i*3-2, column=1).alignment = Alignment(wrap_text=False)
        ws.row_dimensions[i*3-2].height = 30

        img_byte_arr = io.BytesIO()
        image.save(img_byte_arr, format='PNG')
        img_byte_arr = img_byte_arr.getvalue()

        img = XLImage(io.BytesIO(img_byte_arr))
        ws.add_image(img, f'A{i*3-1}')
        ws.row_dimensions[i*3-1].height = 420

    last_row = len(steps) * 3 + 1
    ws.cell(row=last_row, column=1, value=f"Result: {result}")

    if result == "FAIL" and defect:
        ws.cell(row=last_row+1, column=1, value=f"Defect: {defect}")
        ws.cell(row=last_row+1, column=1).alignment = Alignment(wrap_text=False)

    filename = f"{test_name} - {test_number} - {result}.xlsx"
    if export_path:
        filename = os.path.join(export_path, filename)
    wb.save(filename)
    return filename

class TestScriptRunner:
    def __init__(self, master):
        self.master = master
        master.title("Test Script Runner")

        self.test_name = tk.StringVar()
        self.start_number = tk.IntVar()
        self.end_number = tk.IntVar()
        self.current_number = None
        self.steps = []
        self.images = []
        self.result = tk.StringVar(value="PASS")
        self.defect = tk.StringVar()
        self.export_path = None

        self.create_widgets()

    def create_widgets(self):
        ttk.Label(self.master, text="Test Name:").grid(row=0, column=0, sticky="w")
        ttk.Entry(self.master, textvariable=self.test_name).grid(row=0, column=1)

        ttk.Label(self.master, text="Start Number:").grid(row=1, column=0, sticky="w")
        ttk.Entry(self.master, textvariable=self.start_number).grid(row=1, column=1)

        ttk.Label(self.master, text="End Number:").grid(row=2, column=0, sticky="w")
        ttk.Entry(self.master, textvariable=self.end_number).grid(row=2, column=1)

        ttk.Button(self.master, text="Choose Export Location", command=self.choose_export_location).grid(row=3, column=0, columnspan=2)

        ttk.Button(self.master, text="Start Test", command=self.start_test).grid(row=4, column=0, columnspan=2)

        self.steps_text = tk.Text(self.master, height=10, width=50)
        self.steps_text.grid(row=5, column=0, columnspan=2)

        ttk.Button(self.master, text="Add Steps", command=self.add_steps).grid(row=6, column=0)
        ttk.Button(self.master, text="Paste Image (Ctrl+V)", command=self.paste_image).grid(row=6, column=1)

        ttk.Radiobutton(self.master, text="PASS", variable=self.result, value="PASS").grid(row=7, column=0)
        ttk.Radiobutton(self.master, text="FAIL", variable=self.result, value="FAIL").grid(row=7, column=1)

        self.defect_entry = ttk.Entry(self.master, textvariable=self.defect, state="disabled")
        self.defect_entry.grid(row=8, column=0, columnspan=2)

        ttk.Button(self.master, text="Finish Test", command=self.finish_test).grid(row=9, column=0, columnspan=2)

        self.notification_label = ttk.Label(self.master, text="", wraplength=300)
        self.notification_label.grid(row=10, column=0, columnspan=2)

        self.result.trace("w", self.toggle_defect_entry)

    def show_notification(self, message):
        self.notification_label.config(text=message)
        self.master.after(5000, lambda: self.notification_label.config(text=""))

    def choose_export_location(self):
        self.export_path = filedialog.askdirectory()
        if self.export_path:
            self.show_notification(f"Files will be saved to: {self.export_path}")

    def start_test(self):
        if not self.export_path:
            self.show_notification("Please choose an export location first.")
            return
        self.current_number = self.start_number.get()
        self.steps_text.delete("1.0", tk.END)

    def add_steps(self):
        steps_text = self.steps_text.get("1.0", tk.END).strip()
        new_steps = [step.strip() for step in steps_text.split('\n') if step.strip()]
        if new_steps:
            self.steps.extend(new_steps)
            self.steps_text.delete("1.0", tk.END)
            self.show_notification(f"{len(new_steps)} step(s) have been added. Please paste an image for each step.")
        else:
            self.show_notification("Please enter at least one step before adding.")

    def paste_image(self):
        try:
            image = ImageGrab.grabclipboard()
            if image is None:
                raise ValueError("No image found in clipboard")
            
            resized_image = resize_image(image)
            self.images.append(resized_image)
            self.show_notification(f"Image has been pasted and added successfully. ({len(self.images)}/{len(self.steps)})")
            
            if len(self.images) < len(self.steps):
                current_step_number = len(self.images) + 1
                current_step_description = self.steps[current_step_number - 1]
                self.show_notification(f"Step {current_step_number}: {current_step_description}\nPlease paste an image for this step.")
            else:
                self.show_notification("All steps have corresponding images. You can now finish the test.")
        except Exception as e:
            self.show_notification(f"Could not paste image: {str(e)}\nPlease copy an image and try again.")

    def toggle_defect_entry(self, *args):
        if self.result.get() == "FAIL":
            self.defect_entry.config(state="normal")
        else:
            self.defect_entry.config(state="disabled")
            self.defect.set("")

    def finish_test(self):
        if not self.steps or not self.images:
            self.show_notification("Please add steps and paste images before finishing the test.")
            return

        if len(self.steps) != len(self.images):
            self.show_notification("The number of steps and images do not match.")
            return

        try:
            filename = create_excel_report(
                self.test_name.get(),
                self.current_number,
                self.steps,
                self.images,
                self.result.get(),
                self.defect.get() if self.result.get() == "FAIL" else None,
                self.export_path
            )
            self.show_notification(f"Excel report saved as {filename}")

            self.current_number += 1
            if self.current_number <= self.end_number.get():
                self.reset_test()
            else:
                self.master.quit()
        except Exception as e:
            self.show_notification(f"An error occurred while saving the report: {str(e)}")

    def reset_test(self):
        self.steps = []
        self.images = []
        self.result.set("PASS")
        self.defect.set("")
        self.steps_text.delete("1.0", tk.END)

def main():
    root = tk.Tk()
    app = TestScriptRunner(root)
    root.mainloop()

if __name__ == "__main__":
    main()